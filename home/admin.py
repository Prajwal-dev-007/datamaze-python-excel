# admin.py

#from django.contrib import admin
#from .models import ExcelUpload, Person  # Import the ExcelUpload model
"""
from openpyxl import load_workbook
from django.contrib import messages
from django.db import transaction
# admin.py

@admin.register(ExcelUpload)
class ExcelUploadAdmin(admin.ModelAdmin):

    def save_model(self, request, obj, form, change):
        
       # Overriding the save_model method to process the Excel file
        #after it is uploaded via the Django admin interface.
        
        super().save_model(request, obj, form, change)  # Save the Excel file first
        excel_file_path = obj.file.path
        self.process_excel_file(excel_file_path, request)

    def process_excel_file(self, file_path, request):
        
        #Function to process the uploaded Excel file and insert data into PostgreSQL.
        
        wb = load_workbook(file_path)  # Load the workbook
        sheet = wb.active  # Get the active sheet

        with transaction.atomic():  # Ensure atomicity of the operation
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Start at row 2, skipping the header
                if len(row) < 30:
                    messages.error(request, "Invalid row length. Skipping row.")
                    continue

                # Helper function to strip strings safely
                def clean_value(value):
                    if isinstance(value, str):
                        return value.strip()  # Strip spaces if it's a string
                    return value  # Return the value as-is if not a string
                
                def clean_value(value, default_value=None):
                   if isinstance(value, str):
                    return value.strip()
                    return value if value is not None else default_value


                # Clean up the row values and create a new Person record
                Person.objects.create(
                    source=clean_value(row[0] ,"N/A"),
                    Date=clean_value(row[1],"N/A"),
                    client=clean_value(row[2],"N/A"),
                    SPOC=clean_value(row[3],"N/A"),
                    skill=clean_value(row[4],"N/A"),
                    jobcode=clean_value(row[5],"N/A"),
                    Candidatecode=clean_value(row[6],"N/A"),
                    name=clean_value(row[7],"N/A"),
                    contact=clean_value(row[8],"N/A"),
                    email=clean_value(row[9],"N/A"),
                    Location=clean_value(row[10],"N/A"),
                    TEX=clean_value(row[11],"N/A"),
                    REX=clean_value(row[12],"N/A"),
                    Employer=clean_value(row[13],"N/A"),
                    PCTC=clean_value(row[14],"N/A"),
                    ECTC=clean_value(row[15],"N/A"),
                    Notice=clean_value(row[16],"N/A"),
                    IntvDate=clean_value(row[17],"N/A"),
                    IntvTIme=clean_value(row[18],"N/A"),
                    IntvMode=clean_value(row[19],"N/A"),
                    Remarks=clean_value(row[20],"N/A"),
                    ConfirmedProjection=clean_value(row[21], "N/A"),
                    Doj=clean_value(row[22], "N/A"),
                    OfferedCTC=clean_value(row[23], "N/A"),
                    VariableCTC=clean_value(row[24], "N/A"),
                    BillingPrecentage=clean_value(row[25],"N/A"),
                    BillingValue=clean_value(row[26], "N/A"),
                    GrossProfit=clean_value(row[27], "N/A"),
                    NetProfit=clean_value(row[28], "N/A"),
                    RecruterPayout=clean_value(row[29], "N/A")
                )

        messages.success(request, "Excel file processed and data uploaded successfully!")
"""

# admin.py
from django.contrib import admin
from .models import ExcelUpload  # Import the ExcelUpload model
from django.contrib import messages
from .tests import process_excel_file_task  # Import the Celery task

@admin.register(ExcelUpload)
class ExcelUploadAdmin(admin.ModelAdmin):

    def save_model(self, request, obj, form, change):
        """
        Overriding the save_model method to process the Excel file
        after it is uploaded via the Django admin interface.
        """
        super().save_model(request, obj, form, change)  # Save the Excel file first
        excel_file_path = obj.file.path
        
        # Trigger the Celery task to process the file asynchronously
        process_excel_file_task.delay(excel_file_path)

        # Notify the admin user that the task has been triggered
        self.message_user(request, "Excel file uploaded. Processing will start in the background.")
