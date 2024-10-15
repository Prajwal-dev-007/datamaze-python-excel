from django.test import TestCase


# Create your tests here.
# tasks.py
from celery import shared_task
from openpyxl import load_workbook
from django.db import transaction
from .models import Person

import logging

logger = logging.getLogger(__name__)

@shared_task
def process_excel_file_task(file_path):
    import logging
    from openpyxl import load_workbook
    from django.db import transaction
    from .models import Person
    
    logger = logging.getLogger('django')
    logger.info(f"Processing file: {file_path}")

    def clean_value(value, default_value=None):
        if isinstance(value, str):
            return value.strip()
        return value if value is not None else default_value

    try:
        wb = load_workbook(file_path)
        sheet = wb.active

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):

                if all(value is None for value in row):
                    logger.info(f"Stopping processing as all values are None in row: {row}")
                    break  # Stop the loop
                if len(row) < 30:
                    logger.warning(f"Skipping row due to insufficient columns: {row}")
                    continue

                Person.objects.create(
                    source=clean_value(row[0], "N/A"),
                    Date=clean_value(row[1], "N/A"),
                    client=clean_value(row[2], "N/A"),
                    SPOC=clean_value(row[3], "N/A"),
                    skill=clean_value(row[4], "N/A"),
                    jobcode=clean_value(row[5], "N/A"),
                    Candidatecode=clean_value(row[6], "N/A"),
                    name=clean_value(row[7], "N/A"),
                    contact=clean_value(row[8], "N/A"),
                    email=clean_value(row[9], "N/A"),
                    Location=clean_value(row[10], "N/A"),
                    TEX=clean_value(row[11], "N/A"),
                    REX=clean_value(row[12], "N/A"),
                    Employer=clean_value(row[13], "N/A"),
                    PCTC=clean_value(row[14], "N/A"),
                    ECTC=clean_value(row[15], "N/A"),
                    Notice=clean_value(row[16], "N/A"),
                    IntvDate=clean_value(row[17], "N/A"),
                    IntvTIme=clean_value(row[18], "N/A"),
                    IntvMode=clean_value(row[19], "N/A"),
                    Remarks=clean_value(row[20], "N/A"),
                    ConfirmedProjection=clean_value(row[21], "N/A"),
                    Doj=clean_value(row[22], "N/A"),
                    OfferedCTC=clean_value(row[23], "N/A"),
                    VariableCTC=clean_value(row[24], "N/A"),
                    BillingPrecentage=clean_value(row[25], "N/A"),
                    BillingValue=clean_value(row[26], "N/A"),
                    GrossProfit=clean_value(row[27], "N/A"),
                    NetProfit=clean_value(row[28], "N/A"),
                    RecruterPayout=clean_value(row[29], "N/A")
                )
                logger.info(f"Processed row: {row}")

    except Exception as e:
        logger.error(f"Error processing file: {e}")

"""
from celery import shared_task
from openpyxl import load_workbook
from django.db import transaction
from .models import Person
import logging

logger = logging.getLogger(__name__)

@shared_task
def process_excel_file_task(file_path):
    logger.info(f"Processing file: {file_path}")

    def clean_value(value, default_value=None):
        if isinstance(value, str):
            return value.strip()
        return value if value is not None else default_value

    try:
        wb = load_workbook(file_path)
        sheet = wb.active

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if all values in the row are None
                if all(value is None for value in row):
                    logger.info(f"Encountered empty row: {row}. Stopping processing.")
                    break  # Stop processing further rows if you encounter an empty row
                
                if len(row) < 30:
                    logger.warning(f"Skipping row due to insufficient columns: {row}")
                    continue

                try:
                    Person.objects.create(
                        source=clean_value(row[0], "N/A"),
                        Date=clean_value(row[1], "N/A"),
                        client=clean_value(row[2], "N/A"),
                        SPOC=clean_value(row[3], "N/A"),
                        skill=clean_value(row[4], "N/A"),
                        jobcode=clean_value(row[5], "N/A"),
                        Candidatecode=clean_value(row[6], "N/A"),
                        name=clean_value(row[7], "N/A"),
                        contact=clean_value(row[8], "N/A"),
                        email=clean_value(row[9], "N/A"),
                        Location=clean_value(row[10], "N/A"),
                        TEX=clean_value(row[11], "N/A"),
                        REX=clean_value(row[12], "N/A"),
                        Employer=clean_value(row[13], "N/A"),
                        PCTC=clean_value(row[14], "N/A"),
                        ECTC=clean_value(row[15], "N/A"),
                        Notice=clean_value(row[16], "N/A"),
                        IntvDate=clean_value(row[17], "N/A"),
                        IntvTIme=clean_value(row[18], "N/A"),
                        IntvMode=clean_value(row[19], "N/A"),
                        Remarks=clean_value(row[20], "N/A"),
                        ConfirmedProjection=clean_value(row[21], "N/A"),
                        Doj=clean_value(row[22], "N/A"),
                        OfferedCTC=clean_value(row[23], "N/A"),
                        VariableCTC=clean_value(row[24], "N/A"),
                        BillingPrecentage=clean_value(row[25], "N/A"),
                        BillingValue=clean_value(row[26], "N/A"),
                        GrossProfit=clean_value(row[27], "N/A"),
                        NetProfit=clean_value(row[28], "N/A"),
                        RecruterPayout=clean_value(row[29], "N/A")
                    )
                    logger.info(f"Processed row: {row}")
                except Exception as e:
                    logger.error(f"Error processing row {row}: {e}")

    except Exception as e:
        logger.error(f"Error processing file: {e}")

"""






















 
 